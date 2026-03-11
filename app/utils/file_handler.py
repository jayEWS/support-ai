import os
import uuid
import mimetypes
from typing import Optional, Tuple
from app.core.config import settings
from app.core.logging import logger

UPLOAD_DIR = os.path.join("data", "uploads")
CHAT_UPLOAD_DIR = os.path.join(UPLOAD_DIR, "chat")
KNOWLEDGE_UPLOAD_DIR = settings.KNOWLEDGE_DIR

# Allowed extensions by category
ALLOWED_EXTENSIONS = {
    "document": [".pdf", ".txt", ".doc", ".docx", ".xlsx", ".csv"],
    "image": [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"],
    "audio": [".mp3", ".wav", ".ogg", ".m4a", ".aac"],
    "video": [".mp4", ".mov", ".webm", ".avi", ".mkv"],
}

ALL_ALLOWED = [ext for exts in ALLOWED_EXTENSIONS.values() for ext in exts]


def get_file_category(filename: str) -> str:
    """Returns the category of a file based on its extension."""
    ext = os.path.splitext(filename)[1].lower()
    for category, extensions in ALLOWED_EXTENSIONS.items():
        if ext in extensions:
            return category
    return "unknown"


def is_allowed_file(filename: str) -> bool:
    """Check if the file extension is allowed."""
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALL_ALLOWED


# Maximum file sizes (defense-in-depth; routes should also enforce)
MAX_UPLOAD_SIZES = {
    "chat": 10 * 1024 * 1024,       # 10 MB for chat attachments
    "knowledge": 20 * 1024 * 1024,  # 20 MB for knowledge documents
}


def save_upload(file_bytes: bytes, original_filename: str, destination: str = "chat") -> dict:
    """
    Save an uploaded file and return metadata.
    
    Args:
        file_bytes: The raw file bytes
        original_filename: The original filename from the upload
        destination: 'chat' or 'knowledge'
    
    Returns:
        dict with file metadata: filename, path, url, category, size
    
    Raises:
        ValueError: If file exceeds size limit or has disallowed extension
    """
    # P0 Fix: Enforce file size limits
    max_size = MAX_UPLOAD_SIZES.get(destination, MAX_UPLOAD_SIZES["chat"])
    if len(file_bytes) > max_size:
        raise ValueError(
            f"File '{original_filename}' ({len(file_bytes) / (1024*1024):.1f}MB) "
            f"exceeds the {max_size // (1024*1024)}MB limit for {destination} uploads."
        )

    # P0 Fix: Validate file extension
    if not is_allowed_file(original_filename):
        raise ValueError(
            f"File type not allowed: {original_filename}. "
            f"Allowed extensions: {', '.join(ALL_ALLOWED)}"
        )

    if destination == "knowledge":
        upload_dir = KNOWLEDGE_UPLOAD_DIR
    else:
        upload_dir = CHAT_UPLOAD_DIR

    os.makedirs(upload_dir, exist_ok=True)

    # Generate unique filename to avoid collisions
    ext = os.path.splitext(original_filename)[1].lower()
    unique_name = f"{uuid.uuid4().hex[:12]}{ext}"
    file_path = os.path.join(upload_dir, unique_name)

    with open(file_path, "wb") as f:
        f.write(file_bytes)

    category = get_file_category(original_filename)
    mime_type = mimetypes.guess_type(original_filename)[0] or "application/octet-stream"

    metadata = {
        "original_name": original_filename,
        "stored_name": unique_name,
        "path": file_path,
        "url": f"/uploads/chat/{unique_name}" if destination == "chat" else None,
        "category": category,
        "mime_type": mime_type,
        "size": len(file_bytes),
    }

    logger.info(f"File saved: {original_filename} -> {file_path} ({category}, {len(file_bytes)} bytes)")
    return metadata


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from a DOCX file."""
    try:
        from docx import Document
        doc = Document(file_path)
        return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from an XLSX file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        texts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            texts.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception as e:
        logger.error(f"Error extracting text from XLSX: {e}")
        return ""
