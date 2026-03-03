import os
import json
import uuid
import asyncio
import shutil
from fastapi import FastAPI, Request, Depends, HTTPException, BackgroundTasks, UploadFile, File, Form, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import APIKeyHeader, OAuth2PasswordBearer
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter  # NEW: Rate limiting
from slowapi.util import get_remote_address
from contextlib import asynccontextmanager
from typing import AsyncGenerator, Optional, List, Annotated
from datetime import datetime, timezone, timedelta

from app.core.config import settings
from app.core.logging import logger, set_trace_id, LogLatency
from app.core.database import db_manager

def _require_db():
    """Guard: raises 503 if DB is unavailable so app still boots and serves non-DB routes."""
    if db_manager is None:
        raise HTTPException(status_code=503, detail="Database unavailable. Please check DATABASE_URL and DB server connectivity.")
    return db_manager

# ============ RATE LIMITING SETUP ============
limiter = Limiter(key_func=get_remote_address)

# ============ ENVIRONMENT VALIDATION (PRODUCTION SAFETY) ============
def validate_production_config():
    """Validate critical configuration for production deployment."""
    # Minimal critical vars - LLM key can be optional (Groq fallback)
    critical_vars = {
        "AUTH_SECRET_KEY": "Authentication secret key for JWT tokens",
        "DATABASE_URL": "Database connection string",
    }
    
    missing_vars = []
    for var, description in critical_vars.items():
        value = getattr(settings, var, None)
        if not value:
            missing_vars.append(f"  - {var}: {description}")
    
    if missing_vars:
        error_msg = "PRODUCTION CONFIGURATION ERROR - Missing critical environment variables:\n" + "\n".join(missing_vars)
        logger.error(error_msg)
        logger.error("\nWARNING: DEPLOYMENT BLOCKED: Please configure all required environment variables in .env file.")
        raise ValueError(error_msg)
    
    # Warn about optional but recommended settings
    if not settings.OPENAI_API_KEY:
        logger.info("INFO: OPENAI_API_KEY not set. Using Groq fallback for LLM responses.")
    
    # Warn about insecure defaults
    if not settings.COOKIE_SECURE:
        logger.warning("WARNING: COOKIE_SECURE is False. Set to True in production.")
    
    if settings.ALLOWED_ORIGINS == ["*"]:
        logger.warning("WARNING: ALLOWED_ORIGINS is ['*']. Restrict to specific domains in production.")
    
    if settings.MFA_DEV_RETURN_CODE:
        logger.warning("WARNING: MFA_DEV_RETURN_CODE is True. Set to False in production.")
    
    logger.info("[OK] Production configuration validated successfully.")

validate_production_config()

# Services
from app.services.customer_service import CustomerService
from app.services.intent_service import IntentService
from app.services.rag_service import RAGService
from app.services.llm_service import LLMService
from app.services.ticket_service import TicketService
from app.services.escalation_service import EscalationService
from app.services.chat_service import ChatService
from app.services.websocket_manager import manager
from app.webhook.whatsapp import WhatsAppWebhookService

# Background Services
from app.services.sla_service import sla_service
from app.services.routing_service import routing_service

# Schemas
from app.schemas.schemas import IntentType
from app.utils.auth_utils import (
    verify_password,
    create_access_token,
    decode_access_token,
    create_refresh_token,
    hash_token,
    create_mfa_code,
    verify_mfa_code,
    create_mfa_token,
    decode_token,
    create_random_token
)

# Helper functions for Auth
def _extract_bearer_token(request: Request):
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth.split(" ", 1)[1]
    return None

def _set_auth_cookies(response: JSONResponse, access_token: str, refresh_token: str = None):
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        samesite=settings.COOKIE_SAMESITE,
        secure=settings.COOKIE_SECURE,
        max_age=int(settings.ACCESS_TOKEN_EXPIRE_MINUTES * 60)
    )
    if refresh_token:
        response.set_cookie(
            key="refresh_token",
            value=refresh_token,
            httponly=True,
            samesite=settings.COOKIE_SAMESITE,
            secure=settings.COOKIE_SECURE,
            max_age=int(settings.REFRESH_TOKEN_EXPIRE_DAYS * 24 * 60 * 60)
        )

def _clear_auth_cookies(response: JSONResponse):
    response.delete_cookie("access_token")
    response.delete_cookie("refresh_token")

async def get_current_agent(request: Request):
    token = _extract_bearer_token(request) or request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Missing access token")
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    user_id = payload.get("sub")
    agent = _require_db().get_agent(user_id)
    if not agent:
        raise HTTPException(status_code=401, detail="Agent not found")
    return agent

def ensure_utc(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    # Initialize Services with error handling
    try:
        app.state.intent_service = IntentService()
    except Exception as e:
        logger.error(f"Failed to init IntentService: {e}")
        app.state.intent_service = None
    
    try:
        app.state.rag_service = RAGService()
    except Exception as e:
        logger.error(f"Failed to init RAGService: {e}")
        app.state.rag_service = None
    
    try:
        app.state.llm_service = LLMService()
    except Exception as e:
        logger.error(f"Failed to init LLMService: {e}")
        app.state.llm_service = None
    
    try:
        app.state.customer_service = CustomerService()
    except Exception as e:
        logger.error(f"Failed to init CustomerService: {e}")
        app.state.customer_service = None
    
    try:
        app.state.ticket_service = TicketService()
    except Exception as e:
        logger.error(f"Failed to init TicketService: {e}")
        app.state.ticket_service = None
    
    try:
        app.state.escalation_service = EscalationService()
    except Exception as e:
        logger.error(f"Failed to init EscalationService: {e}")
        app.state.escalation_service = None
    
    # Initialize GCS Service (Phase 2)
    try:
        from app.services.gcs_service import init_gcs_service
        app.state.gcs_service = init_gcs_service()
    except Exception as e:
        logger.warning(f"GCS Service init skipped: {e}")
        app.state.gcs_service = None
    
    try:
        app.state.chat_service = ChatService(app.state.rag_service if app.state.rag_service else None)
    except Exception as e:
        logger.error(f"Failed to init ChatService: {e}")
        app.state.chat_service = None
    
    # Start background workers - TEMPORARILY DISABLED for stability
    # try:
    #     asyncio.create_task(sla_service.monitor_breaches())
    #     asyncio.create_task(routing_service.process_queue())
    # except Exception as e:
    #     logger.error(f"Failed to start background workers: {e}")
    
    logger.info("All services and background workers initialized.")
    yield
    # Cleanup

app = FastAPI(title="Support Portal Edgeworks", lifespan=lifespan)

# Add CORS Middleware - MUST be configured correctly for production
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS if settings.ALLOWED_ORIGINS else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add rate limiter to app
app.state.limiter = limiter

# Setup Templates & Static Files
templates = Jinja2Templates(directory="templates")
os.makedirs(os.path.join("data", "uploads", "chat"), exist_ok=True)
app.mount("/uploads", StaticFiles(directory=os.path.join("data", "uploads")), name="uploads")

# ============ UI Routes ============

@app.get("/", response_class=HTMLResponse)
async def portal_dashboard(request: Request):
    response = templates.TemplateResponse(request, "index.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

@app.get("/chat", response_class=HTMLResponse)
async def live_chat_demo(request: Request):
    return templates.TemplateResponse(request, "chat.html")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    response = templates.TemplateResponse(request, "login.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    return templates.TemplateResponse(request, "admin.html")

# ============ Auth APIs ============

@app.post("/api/auth/login")
@limiter.limit("5/minute")  # Prevent brute force
async def login(request: Request):
    try:
        data = await request.json()
        email, password = data.get("email"), data.get("password")
        if not email or not email.lower().endswith("@edgeworks.com.sg"):
            raise HTTPException(status_code=403, detail="Only @edgeworks.com.sg accounts are allowed")
            
        db = _require_db()
        agent = db.get_agent_by_email(email)
        if not agent or not verify_password(password, agent.get("hashed_password")):
            raise HTTPException(status_code=401, detail="Incorrect email or password")
        
        # Simplified: Auto-login for now (MFA logic can be added back if needed)
        access_token = create_access_token(data={"sub": agent["user_id"], "role": agent["role"]})
        refresh_token = create_refresh_token()
        refresh_hash = hash_token(refresh_token)
        refresh_expires = datetime.now(timezone.utc) + timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
        db.create_refresh_token(agent["user_id"], refresh_hash, refresh_expires, request.headers.get("user-agent"))
        
        db.log_action(agent["user_id"], "login", "agent", agent["user_id"])

        response = JSONResponse({
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": agent["role"],
            "name": agent.get("name", agent["user_id"])
        })
        _set_auth_cookies(response, access_token, refresh_token)
        return response
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/api/auth/me")
async def get_me(agent: Annotated[dict, Depends(get_current_agent)]):
    return agent

@app.post("/api/auth/logout")
async def logout(request: Request):
    incoming = request.cookies.get("refresh_token")
    if incoming and db_manager:
        db_manager.revoke_refresh_token(hash_token(incoming))
    response = JSONResponse({"status": "success"})
    _clear_auth_cookies(response)
    return response

@app.post("/api/auth/refresh")
async def refresh_token_route(request: Request):
    incoming = request.cookies.get("refresh_token")
    if not incoming: raise HTTPException(status_code=401, detail="Missing refresh token")
    db = _require_db()
    token_hash = hash_token(incoming)
    stored = db.get_refresh_token(token_hash)
    if not stored or stored.get("revoked_at"):
        raise HTTPException(status_code=401, detail="Invalid refresh token")

    agent = db.get_agent(stored["user_id"])
    if not agent: raise HTTPException(status_code=401, detail="Agent not found")

    db.revoke_refresh_token(token_hash)
    new_refresh = create_refresh_token()
    refresh_expires = datetime.now(timezone.utc) + timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
    db.create_refresh_token(agent["user_id"], hash_token(new_refresh), refresh_expires, request.headers.get("user-agent"))

    access_token = create_access_token(data={"sub": agent["user_id"], "role": agent["role"]})
    response = JSONResponse({"access_token": access_token, "token_type": "bearer", "role": agent["role"]})
    _set_auth_cookies(response, access_token, new_refresh)
    return response

# ============ GOOGLE OAUTH LOGIN ============
@app.get("/api/auth/google")
@limiter.limit("10/minute")
async def google_oauth_redirect(request: Request):
    """Redirect user to Google OAuth consent screen"""
    if not settings.GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=501, detail="Google OAuth not configured. Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET.")
    
    import urllib.parse, secrets
    state = secrets.token_urlsafe(16)
    # Store state in session cookie for CSRF protection
    params = {
        "client_id": settings.GOOGLE_CLIENT_ID,
        "redirect_uri": settings.GOOGLE_REDIRECT_URI,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "offline",
        "state": state,
        "prompt": "select_account",
        "hd": "edgeworks.com.sg",
    }
    google_auth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
    response = RedirectResponse(url=google_auth_url)
    response.set_cookie("oauth_state", state, max_age=300, httponly=True, samesite="lax")
    return response


@app.get("/api/auth/google/callback")
@limiter.limit("10/minute")
async def google_oauth_callback(request: Request, code: str = None, state: str = None, error: str = None):
    """Handle Google OAuth callback, exchange code for tokens, log user in"""
    import urllib.parse as _up
    import httpx as _httpx

    if error:
        return RedirectResponse(url=f"/login?error={_up.quote(error)}")

    if not code:
        return RedirectResponse(url="/login?error=missing_code")

    # Validate state for CSRF protection
    stored_state = request.cookies.get("oauth_state")
    if not stored_state or stored_state != state:
        return RedirectResponse(url="/login?error=invalid_state")

    try:
        # Exchange code for tokens
        async with _httpx.AsyncClient() as client:
            token_resp = await client.post(
                "https://oauth2.googleapis.com/token",
                data={
                    "code": code,
                    "client_id": settings.GOOGLE_CLIENT_ID,
                    "client_secret": settings.GOOGLE_CLIENT_SECRET,
                    "redirect_uri": settings.GOOGLE_REDIRECT_URI,
                    "grant_type": "authorization_code",
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                timeout=10,
            )
        if token_resp.status_code != 200:
            logger.error(f"Google token exchange failed: {token_resp.text}")
            return RedirectResponse(url="/login?error=token_exchange_failed")

        token_data = token_resp.json()
        id_token_str = token_data.get("id_token")

        # Verify the ID token with Google
        from google.oauth2 import id_token as google_id_token_lib
        from google.auth.transport import requests as google_requests
        try:
            id_info = google_id_token_lib.verify_oauth2_token(
                id_token_str,
                google_requests.Request(),
                settings.GOOGLE_CLIENT_ID,
            )
        except Exception as verify_err:
            logger.error(f"Google ID token verification failed: {verify_err}")
            return RedirectResponse(url="/login?error=invalid_token")

        google_sub = id_info["sub"]
        email = id_info.get("email", "")
        if not email or not email.lower().endswith("@edgeworks.com.sg"):
            logger.warning(f"Google login domain mismatch: {email}")
            return RedirectResponse(url="/login?error=invalid_domain")
            
        name = id_info.get("name", email.split("@")[0] if email else "Agent")

        db = _require_db()
        agent = db.get_agent_by_google_id(google_sub)
        
        # Designated System Admins
        system_admins = ["support@edgeworks.com.sg", "jay@edgeworks.com.sg"]
        is_sys_admin = email.lower() in system_admins

        if not agent:
            # Check if an agent with this email already exists
            agent_data = db.get_agent_by_email(email)
            if agent_data:
                # Link existing agent
                db.update_agent_auth(agent_data["user_id"], google_id=google_sub)
                agent = db.get_agent(agent_data["user_id"])
            else:
                # STRICT: Only auto-create designate admins via Google. 
                # Others must use Magic Link (Token verification) first.
                if is_sys_admin:
                    agent = db.create_or_get_agent(
                        user_id=f"google_{uuid.uuid4().hex[:16]}",
                        name=name,
                        email=email,
                        department="Management",
                        google_id=google_sub,
                    )
                else:
                    logger.warning(f"Unauthorized Google registration attempt: {email}")
                    return RedirectResponse(url="/login?error=use_magic_link")
        else:
            db.update_agent_auth(agent["user_id"], google_id=google_sub)
            
        # Ensure System Admin role for designated emails
        if is_sys_admin:
            db.create_role("System Admin", "Full system access")
            db.assign_role_to_agent(agent["user_id"], "System Admin")
            agent = db.get_agent(agent["user_id"])

        access_token = create_access_token(data={"sub": agent["user_id"], "role": agent.get("role", "agent")})
        refresh_token = create_refresh_token()
        refresh_hash = hash_token(refresh_token)
        refresh_expires = datetime.now(timezone.utc) + timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
        db.create_refresh_token(agent["user_id"], refresh_hash, refresh_expires, request.headers.get("user-agent"))
        db.log_action(agent["user_id"], "login_google", "agent", agent["user_id"])

        # Redirect to login page so JS can store token and go to /admin
        qs = _up.urlencode({
            "access_token": access_token,
            "role": agent.get("role", "agent"),
            "name": agent.get("name", agent["user_id"]),
        })
        response = RedirectResponse(url=f"/login?oauth_success=1&{qs}")
        response.delete_cookie("oauth_state")
        _set_auth_cookies(response, access_token, refresh_token)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Google OAuth callback error: {e}")
        return RedirectResponse(url="/login?error=oauth_failed")


@app.post("/api/auth/google/login")
@limiter.limit("10/minute")
async def google_login_token(request: Request):
    """Handle Google Login via ID token (from GIS/One Tap)"""
    try:
        data = await request.json()
        id_token_str = data.get("id_token")
        if not id_token_str:
            raise HTTPException(status_code=400, detail="Missing ID token")

        # Verify the ID token with Google
        from google.oauth2 import id_token as google_id_token_lib
        from google.auth.transport import requests as google_requests
        
        try:
            id_info = google_id_token_lib.verify_oauth2_token(
                id_token_str,
                google_requests.Request(),
                settings.GOOGLE_CLIENT_ID,
            )
        except Exception as verify_err:
            logger.error(f"Google ID token verification failed: {verify_err}")
            raise HTTPException(status_code=401, detail="Invalid Google token")

        google_sub = id_info["sub"]
        email = id_info.get("email", "")
        if not email or not email.lower().endswith("@edgeworks.com.sg"):
            logger.warning(f"Google token domain mismatch: {email}")
            raise HTTPException(status_code=403, detail="Only @edgeworks.com.sg accounts are allowed")

        name = id_info.get("name", email.split("@")[0] if email else "Agent")

        db = _require_db()
        agent = db.get_agent_by_google_id(google_sub)

        # Designated System Admins
        system_admins = ["support@edgeworks.com.sg", "jay@edgeworks.com.sg"]
        is_sys_admin = email.lower() in system_admins

        if not agent:
            # Check if an agent with this email already exists
            agent_data = db.get_agent_by_email(email)
            if agent_data:
                # Link existing agent
                db.update_agent_auth(agent_data["user_id"], google_id=google_sub)
                agent = db.get_agent(agent_data["user_id"])
            else:
                # STRICT: Only designate admins can register via Token Login.
                # Others must use Magic Link flow first.
                if is_sys_admin:
                    agent = db.create_or_get_agent(
                        user_id=f"google_{uuid.uuid4().hex[:16]}",
                        name=name,
                        email=email,
                        department="Management",
                        google_id=google_sub,
                    )
                else:
                    raise HTTPException(status_code=403, detail="First-time registration must use Magic Link verification.")
        else:
            db.update_agent_auth(agent["user_id"], google_id=google_sub)

        # Ensure System Admin role for designated emails
        if is_sys_admin:
            db.create_role("System Admin", "Full system access")
            db.assign_role_to_agent(agent["user_id"], "System Admin")
            agent = db.get_agent(agent["user_id"])

        access_token = create_access_token(data={"sub": agent["user_id"], "role": agent.get("role", "agent")})
        refresh_token = create_refresh_token()
        refresh_hash = hash_token(refresh_token)
        refresh_expires = datetime.now(timezone.utc) + timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
        db.create_refresh_token(agent["user_id"], refresh_hash, refresh_expires, request.headers.get("user-agent"))
        db.log_action(agent["user_id"], "login_google_token", "agent", agent["user_id"])

        response = JSONResponse({
            "access_token": access_token,
            "token_type": "bearer",
            "role": agent.get("role", "agent"),
            "name": agent.get("name", agent["user_id"])
        })
        _set_auth_cookies(response, access_token, refresh_token)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Google token login error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during Google login")

# ============ MAGIC LINK LOGIN ============
@app.post("/api/auth/magic-link/request")
@limiter.limit("3/minute")
async def request_magic_link(request: Request, background_tasks: BackgroundTasks):
    """Request a magic link login"""
    try:
        data = await request.json()
        email = data.get("email")
        if not email or not email.lower().endswith("@edgeworks.com.sg"):
            raise HTTPException(status_code=403, detail="Only @edgeworks.com.sg accounts are allowed")
        
        if not email:
            raise HTTPException(status_code=400, detail="Email required")
        
        # Generate magic link token
        magic_token = create_random_token()
        magic_hash = hash_token(magic_token)
        
        # Store magic link with 15 minute expiry
        magic_expires = datetime.now(timezone.utc) + timedelta(minutes=15)
        _require_db().create_magic_link(email, magic_hash, magic_expires)
        
        # Build the magic link URL (for production, this should be from settings.BASE_URL)
        magic_link_url = f"{settings.BASE_URL}/api/auth/magic-link/verify?token={magic_token}&email={email}"
        
        # Send email in background
        from app.utils.email_utils import send_magic_link_email
        background_tasks.add_task(send_magic_link_email, email, magic_link_url)
        
        # Log the attempt
        logger.info(f"Magic link requested for: {email}")
        
        return JSONResponse({
            "status": "success",
            "message": "Check your email for the magic link. If you don't see it, check your spam folder."
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Magic link request error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create magic link")

@app.get("/api/auth/magic-link/verify")
async def verify_magic_link(token: str, email: str, request: Request):
    """Verify magic link and log in user"""
    import urllib.parse as _up
    try:
        # Check magic link validity
        magic_hash = hash_token(token)
        db = _require_db()
        magic_link = db.get_magic_link(email, magic_hash)
        
        if not magic_link:
            raise HTTPException(status_code=401, detail="Invalid or expired magic link")
        
        # Handle both timezone-aware and naive datetimes
        expires_at = magic_link.get("expires_at")
        now = datetime.now(timezone.utc)
        
        # Ensure both are timezone-aware for comparison
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        
        if expires_at < now:
            db.revoke_magic_link(email, magic_hash)
            raise HTTPException(status_code=401, detail="Magic link expired")
        
        # Get or create agent
        agent = db.get_agent_by_email(email)
        if not agent:
            agent = db.create_or_get_agent(
                user_id=f"ml_{uuid.uuid4().hex[:16]}",
                name=email.split("@")[0],
                email=email,
                department="Support"
            )
        
        # Revoke used magic link
        db.revoke_magic_link(email, magic_hash)
        
        # Create session tokens
        access_token = create_access_token(data={"sub": agent["user_id"], "role": agent.get("role", "agent")})
        refresh_token = create_refresh_token()
        refresh_hash = hash_token(refresh_token)
        refresh_expires = datetime.now(timezone.utc) + timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
        db.create_refresh_token(agent["user_id"], refresh_hash, refresh_expires, request.headers.get("user-agent"))
        
        db.log_action(agent["user_id"], "login_magic_link", "agent", agent["user_id"])
        
        # Redirect to login page with token in URL so JS can store it and go to /admin
        qs = _up.urlencode({
            "access_token": access_token,
            "role": agent.get("role", "agent"),
            "name": agent.get("name", agent["user_id"]),
        })
        response = RedirectResponse(url=f"/login?magic_success=1&{qs}", status_code=302)
        _set_auth_cookies(response, access_token, refresh_token)
        return response
    except HTTPException as he:
        return RedirectResponse(url=f"/login?error={_up.quote(he.detail)}", status_code=302)
    except Exception as e:
        logger.error(f"Magic link verification error: {e}")
        return RedirectResponse(url="/login?error=verification_failed", status_code=302)

# ============ Management APIs ============

@app.get("/api/tickets")
async def get_tickets(agent: Annotated[dict, Depends(get_current_agent)], filter: str = 'all'):
    return _require_db().get_all_tickets(filter_type=filter)

@app.get("/api/tickets/counts")
async def get_ticket_counts(agent: Annotated[dict, Depends(get_current_agent)]):
    return _require_db().get_ticket_counts()

@app.get("/api/tickets/{ticket_id}/history")
async def get_ticket_history(ticket_id: int, agent: Annotated[dict, Depends(get_current_agent)]):
    db = _require_db()
    ticket = db.get_session().query(db.engine).get(ticket_id) if False else None
    # Use db method directly instead of raw query
    history = db.get_unified_history(None, ticket_id)
    return {"history": history}

@app.patch("/api/tickets/{ticket_id}/status")
async def update_ticket_status(ticket_id: int, data: dict, agent: Annotated[dict, Depends(get_current_agent)]):
    if "status" in data:
        _require_db().update_ticket_status(ticket_id, data["status"])
    return {"status": "success"}

@app.get("/api/analytics")
async def get_analytics(request: Request, agent: Annotated[dict, Depends(get_current_agent)]):
    # Mock/Actual Analytics integration
    metrics = _require_db().get_ticket_metrics()
    return {
        "metrics": metrics,
        "agent_performance": [],
        "priority_distribution": {"High": 5, "Medium": 10, "Low": 2},
        "volume_trends": [],
        "ai_trends": "Tren AI sedang dianalisis..."
    }

@app.get("/api/agents")
async def get_agents(agent: Annotated[dict, Depends(get_current_agent)]):
    return _require_db().get_all_agents()

@app.get("/api/knowledge")
async def get_knowledge(agent: Annotated[dict, Depends(get_current_agent)]):
    return _require_db().get_all_knowledge()

@app.post("/api/knowledge/upload")
async def upload_knowledge(
    agent: Annotated[dict, Depends(get_current_agent)],
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...)
):
    for file in files:
        file_path = os.path.join(settings.KNOWLEDGE_DIR, file.filename)
        with open(file_path, "wb") as buffer: shutil.copyfileobj(file.file, buffer)
        _require_db().save_knowledge_metadata(filename=file.filename, file_path=file_path, uploaded_by=agent["user_id"])
        
        # Phase 2: Upload to GCS
        try:
            from app.services.gcs_service import get_gcs_service
            gcs = get_gcs_service()
            if gcs.enabled:
                gcs.upload_file(file_path, file.filename)
        except Exception as gcs_err:
            logger.warning(f"[GCS] Upload sync failed for {file.filename}: {gcs_err}")
    
    # Trigger re-indexing
    background_tasks.add_task(_reindex_knowledge)
    return {"status": "success"}

async def _reindex_knowledge():
    """Re-index knowledge base: reload vector store and reinitialize hybrid search."""
    try:
        rag_svc = app.state.rag_service
        if rag_svc:
            rag_svc.vector_store = rag_svc._load_vector_store()
            rag_svc._initialize_hybrid_search()
            # Update all knowledge statuses to 'Indexed'
            all_kb = _require_db().get_all_knowledge()
            for kb in all_kb:
                _require_db().update_knowledge_status(kb['filename'], 'Indexed')
            logger.info(f"[OK] Knowledge base re-indexed: {len(rag_svc.all_documents)} chunks")
    except Exception as e:
        logger.error(f"Re-index failed: {e}")

@app.post("/api/knowledge/reindex")
async def reindex_knowledge(
    agent: Annotated[dict, Depends(get_current_agent)],
    background_tasks: BackgroundTasks
):
    """Manually trigger knowledge base re-indexing (Train AI)"""
    background_tasks.add_task(_reindex_knowledge)
    return {"status": "reindexing", "message": "Knowledge base re-indexing started. This may take a moment."}

@app.get("/api/knowledge/stats")
async def knowledge_stats(agent: Annotated[dict, Depends(get_current_agent)]):
    """Get knowledge base statistics"""
    try:
        all_kb = _require_db().get_all_knowledge()
        rag_svc = app.state.rag_service
        total_files = len(all_kb)
        indexed_count = sum(1 for k in all_kb if k.get('status') == 'Indexed')
        total_chunks = len(rag_svc.all_documents) if rag_svc and rag_svc.all_documents else 0
        has_vector_store = bool(rag_svc and rag_svc.vector_store)
        last_upload = all_kb[0]['upload_date'] if all_kb and all_kb[0].get('upload_date') else None
        return {
            "total_files": total_files,
            "indexed_files": indexed_count,
            "total_chunks": total_chunks,
            "vector_store_ready": has_vector_store,
            "last_upload": last_upload
        }
    except Exception as e:
        logger.error(f"Knowledge stats error: {e}")
        return {"total_files": 0, "indexed_files": 0, "total_chunks": 0, "vector_store_ready": False}

@app.post("/api/knowledge/ingest-url")
async def ingest_knowledge_from_url(
    agent: Annotated[dict, Depends(get_current_agent)],
    request: Request
):
    """Ingest knowledge base from URL"""
    try:
        data = await request.json()
        url = data.get("url")
        
        if not url:
            return JSONResponse({"error": "URL is required"}, status_code=400)
        
        # Use RAG engine to ingest from URL
        try:
            from app.services.rag_engine import rag_engine
            if not rag_engine:
                return JSONResponse({"error": "RAG Engine not initialized"}, status_code=500)
            
            filename = await rag_engine.ingest_from_url(url, uploaded_by=agent["user_id"])
            return JSONResponse({"status": "success", "filename": filename, "message": f"Content from {url} ingested successfully"})
        except Exception as import_err:
            logger.error(f"RAG engine import error: {str(import_err)}")
            raise
    except ValueError as e:
        return JSONResponse({"error": str(e)}, status_code=400)
    except Exception as e:
        logger.error(f"URL ingestion error: {str(e)}")
        return JSONResponse({"error": f"Failed to ingest from URL: {str(e)}"}, status_code=500)

@app.delete("/api/knowledge/{filename}")
async def delete_knowledge(
    filename: str,
    agent: Annotated[dict, Depends(get_current_agent)]
):
    try:
        from app.services.rag_engine import rag_engine
        if not rag_engine:
            return JSONResponse({"error": "RAG Engine not initialized"}, status_code=500)
        
        rag_engine.delete_knowledge_document(filename)
        return {"status": "success", "message": f"Deleted {filename}"}
    except Exception as e:
        logger.error(f"Error deleting knowledge: {str(e)}")
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/api/knowledge/batch-delete")
async def batch_delete_knowledge(
    request: Request,
    agent: Annotated[dict, Depends(get_current_agent)]
):
    try:
        data = await request.json()
        filenames = data.get("filenames", [])
        
        if not filenames:
            return JSONResponse({"error": "No filenames provided"}, status_code=400)
            
        from app.services.rag_engine import rag_engine
        if not rag_engine:
            return JSONResponse({"error": "RAG Engine not initialized"}, status_code=500)
            
        rag_engine.delete_knowledge_documents(filenames)
        return {"status": "success", "message": f"Deleted {len(filenames)} files"}
    except Exception as e:
        logger.error(f"Error batch deleting knowledge: {str(e)}")
        return JSONResponse({"error": str(e)}, status_code=500)

# ============ Customer Management APIs ============

@app.get("/api/customers")
async def get_customers(agent: Annotated[dict, Depends(get_current_agent)]):
    """List all customers (users) from DB."""
    return _require_db().get_all_users()

@app.get("/api/customers/{identifier}")
async def get_customer(identifier: str, agent: Annotated[dict, Depends(get_current_agent)]):
    """Get single customer details."""
    user = _require_db().get_user(identifier)
    if not user:
        return JSONResponse({"error": "Customer not found"}, status_code=404)
    return user

@app.post("/api/customers")
async def create_customer(request: Request, agent: Annotated[dict, Depends(get_current_agent)]):
    """Create or update a single customer."""
    data = await request.json()
    identifier = data.get("identifier") or data.get("phone") or data.get("id")
    name = data.get("name")
    company = data.get("company", "")
    outlet = data.get("outlet") or data.get("outlet_pos") or company
    position = data.get("position", "")
    if not identifier or not name:
        return JSONResponse({"error": "identifier and name are required"}, status_code=400)
    _require_db().create_or_update_user(identifier, name=name, company=company, position=position, outlet_pos=outlet, state="complete")
    return {"status": "success", "identifier": identifier, "name": name}

@app.post("/api/customers/import")
async def import_customers(
    agent: Annotated[dict, Depends(get_current_agent)],
    file: UploadFile = File(...)
):
    """Bulk import customers from CSV/Excel file.
    Expected columns: name, company/outlet, phone/identifier, position (optional)
    """
    import csv
    import io

    filename = file.filename.lower()
    content = await file.read()
    rows = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)})
                wb.close()
            except ImportError:
                return JSONResponse({"error": "openpyxl package not installed. Use CSV format or install openpyxl."}, status_code=400)
        else:
            return JSONResponse({"error": "Unsupported format. Use CSV or XLSX."}, status_code=400)
    except Exception as e:
        return JSONResponse({"error": f"Failed to parse file: {str(e)}"}, status_code=400)

    if not rows:
        return JSONResponse({"error": "File is empty or has no data rows"}, status_code=400)

    imported = 0
    skipped = 0
    errors_list = []

    for i, row in enumerate(rows, 1):
        # Normalize column names (support various naming conventions)
        r = {k.strip().lower().replace(" ", "_"): v for k, v in row.items()}
        name = r.get("name") or r.get("nama") or r.get("customer_name") or r.get("full_name", "")
        company = r.get("company") or r.get("perusahaan") or r.get("outlet") or r.get("outlet_pos") or r.get("toko", "")
        identifier = r.get("identifier") or r.get("phone") or r.get("id") or r.get("nomor") or r.get("telepon") or r.get("whatsapp", "")
        position = r.get("position") or r.get("jabatan", "")

        if not name and not identifier:
            skipped += 1
            continue

        # Auto-generate identifier if missing
        if not identifier:
            identifier = f"imp_{name.lower().replace(' ', '_')}_{i}"

        try:
            _require_db().create_or_update_user(
                identifier=identifier.strip(),
                name=name.strip(),
                company=company.strip(),
                position=position.strip(),
                outlet_pos=(company or "").strip(),
                state="complete"
            )
            imported += 1
        except Exception as e:
            errors_list.append(f"Row {i}: {str(e)}")
            skipped += 1

    return {
        "status": "success",
        "imported": imported,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors_list[:10]  # Return first 10 errors max
    }

@app.delete("/api/customers/{identifier}")
async def delete_customer(identifier: str, agent: Annotated[dict, Depends(get_current_agent)]):
    """Delete a customer record."""
    db = _require_db()
    user = db.get_user(identifier)
    if not user:
        return JSONResponse({"error": "Customer not found"}, status_code=404)
    try:
        session = db.get_session()
        from app.models.models import User
        u = session.query(User).get(identifier)
        if u:
            session.delete(u)
            session.commit()
        return {"status": "success", "message": f"Deleted customer {identifier}"}
    except Exception as e:
        logger.error(f"Delete customer error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        db.Session.remove()

# ============ GCS (Google Cloud Storage) APIs ============

@app.get("/api/gcs/status")
async def gcs_status(agent: Annotated[dict, Depends(get_current_agent)]):
    """Get GCS service status and file count."""
    try:
        from app.services.gcs_service import get_gcs_service
        gcs = get_gcs_service()
        return gcs.get_status()
    except Exception as e:
        return {"enabled": False, "error": str(e)}

@app.post("/api/gcs/sync")
async def gcs_sync(agent: Annotated[dict, Depends(get_current_agent)]):
    """Sync all local knowledge files to GCS bucket."""
    try:
        from app.services.gcs_service import get_gcs_service
        gcs = get_gcs_service()
        if not gcs.enabled:
            return JSONResponse({"error": "GCS is not enabled. Set GCS_ENABLED=True in .env"}, status_code=400)
        
        results = await gcs.async_sync_local_to_gcs()
        synced = sum(1 for v in results.values() if v != "FAILED" and not str(v).startswith("_"))
        failed = sum(1 for v in results.values() if v == "FAILED")
        return {
            "status": "success",
            "synced": synced,
            "failed": failed,
            "details": results
        }
    except Exception as e:
        logger.error(f"GCS sync error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/gcs/files")
async def gcs_list_files(agent: Annotated[dict, Depends(get_current_agent)]):
    """List all files in GCS knowledge bucket."""
    try:
        from app.services.gcs_service import get_gcs_service
        gcs = get_gcs_service()
        if not gcs.enabled:
            return JSONResponse({"error": "GCS is not enabled"}, status_code=400)
        
        files = gcs.list_files()
        return {"files": files, "count": len(files)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/macros")
async def get_macros(agent: Annotated[dict, Depends(get_current_agent)]):
    return _require_db().get_macros()

@app.get("/api/audit-logs")
async def get_audit_logs(agent: Annotated[dict, Depends(get_current_agent)]):
    return _require_db().get_audit_logs()

# ============ Public / Portal APIs ============

@app.get("/api/history")
@limiter.limit("10/minute")
async def get_chat_history(request: Request, user_id: str = "web_portal_user"):
    return {"history": _require_db().get_messages(user_id)}

@app.get("/api/history/sessions")
@limiter.limit("10/minute")
async def get_history_sessions(request: Request, user_id: str = "web_portal_user"):
    return {"sessions": _require_db().get_unified_history(user_id)}

@app.post("/api/chat")
@limiter.limit("5/minute")
async def chat_directly(
    request: Request,
    message: Optional[str] = Form(None),
    user_id: Optional[str] = Form(None),
    language: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    set_trace_id()
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        try:
            data = await request.json()
            message, user_id, language = data.get("message"), data.get("user_id", "web_portal_user"), data.get("language")
        except: return JSONResponse({"error": "Invalid JSON"}, status_code=400)
    
    response_data, status_code = await app.state.chat_service.process_portal_message(
        query=message, user_id=user_id or "web_portal_user", file=file, language=language
    )
    return JSONResponse(response_data, status_code=status_code)

@app.post("/api/close-session")
async def close_session(request: Request):
    data = await request.json()
    user_id = data.get("user_id", "web_portal_user")
    option = data.get("option", "close")  # "close" | "ticket" | "ticket_and_notify"
    
    # Support legacy numeric options
    if option == 1 or option == "1":
        option = "ticket_and_notify"
    elif option == 2 or option == "2":
        option = "close"
    
    result = await app.state.chat_service.close_chat(user_id, option)
    return result

# ============ WhatsApp Webhook ============

@app.post("/webhook/whatsapp")
@limiter.limit("10/minute")  # NEW: Rate limit to prevent abuse
async def whatsapp_webhook(request: Request, background_tasks: BackgroundTasks):
    trace_id = set_trace_id()
    message = await WhatsAppWebhookService.normalize_payload(request)
    customer = await app.state.customer_service.get_or_register_customer(message.sender)
    classification = await app.state.intent_service.classify(message.text)
    
    response_text = ""
    if classification.intent == IntentType.CRITICAL:
        response_text = await app.state.escalation_service.escalate(customer.identifier, f"CRITICAL: {message.text}", message.text, True)
    elif classification.intent == IntentType.ESCALATION:
        response_text = await app.state.escalation_service.escalate(customer.identifier, f"Escalation: {message.text}", message.text)
    elif classification.intent == IntentType.DEEP_REASONING:
        response_text = await app.state.llm_service.reason(message.text)
    else: 
        rag_res = await app.state.rag_service.query(message.text)
        if rag_res.confidence < 0.5:
            response_text = await app.state.escalation_service.escalate(customer.identifier, f"Low Confidence", message.text)
        else: response_text = rag_res.answer

    if customer.is_new:
        response_text = f"{app.state.customer_service.get_personalized_greeting(customer)}\n\n{response_text}"

    # Actually Send the reply back to WhatsApp via Bird API
    try:
        from app.adapters.whatsapp_bird import send_whatsapp_message
        await send_whatsapp_message(message.sender, response_text)
    except Exception as e:
        logger.error(f"Failed to send WhatsApp message back: {e}")

    return {"status": "success", "trace_id": trace_id, "response": response_text}

# ============ WebSocket ============

@app.websocket("/ws/chat/{session_id}")
async def websocket_chat_endpoint(websocket: WebSocket, session_id: int, user_id: str, user_type: str):
    try:
        db = _require_db() if db_manager else None
        if not db:
            await websocket.close(code=1008)
            return
        session = db.get_chat_session(session_id)
        if not session:
            await websocket.close(code=1008)
            return
        await manager.connect(session_id, user_id, user_type, websocket)
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)
            if msg.get("event") == "message":
                content = msg.get("content")
                msg_id = db.save_chat_message(session_id, user_id, user_type, content)
                await manager.broadcast(session_id, {"event": "message", "message_id": msg_id, "sender_id": user_id, "sender_type": user_type, "content": content, "sent_at": datetime.now(timezone.utc).isoformat()})
    except: manager.disconnect(session_id, websocket)

@app.get("/health")
async def health():
    health_data = {"status": "ok"}
    
    # Phase 2: Include GCS status
    try:
        from app.services.gcs_service import get_gcs_service
        gcs = get_gcs_service()
        health_data["gcs"] = {"enabled": gcs.enabled, "bucket": settings.GCS_BUCKET_NAME if gcs.enabled else None}
    except Exception:
        health_data["gcs"] = {"enabled": False}
    
    return health_data

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8001, reload=True)
